"""Accounts Dashboard service — broker totals, per-user PNL breakdown, exports.

Reuses the PNL calculation pattern from pnl_sharing_service / admin_settlement_service
but operates independently (no PnlSharingAgreement required).
"""

from __future__ import annotations

import asyncio
import logging
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from beanie import PydanticObjectId
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.models.pnl_sharing import AgreementStatus, AgreementType, PnlSharingAgreement
from app.models.position import Position, PositionStatus
from app.models.trade import Trade
from app.models.transaction import TransactionType, WalletTransaction
from app.models.user import User, UserRole, UserStatus
from app.models.wallet import Wallet
from app.services import market_data_service
from app.services.admin_settlement_service import _realised_inr
from app.utils.decimal_utils import quantize_money, to_decimal
from app.utils.time_utils import IST

logger = logging.getLogger(__name__)

TRADING_ROLES = [UserRole.CLIENT.value, UserRole.DEALER.value, UserRole.MASTER.value]


# ── Week options for dropdown ────────────────────────────────────────

def generate_week_options(num_weeks: int = 16) -> list[dict[str, str]]:
    """Last N IST weeks (Monday–Sunday). Most recent first."""
    today = datetime.now(IST).date()
    monday = today - timedelta(days=today.weekday())
    weeks = []
    for i in range(num_weeks):
        w_mon = monday - timedelta(weeks=i)
        w_sun = w_mon + timedelta(days=6)
        weeks.append({
            "label": f"Week_{w_mon.isoformat()}",
            "start": w_mon.isoformat(),
            "end": w_sun.isoformat(),
        })
    return weeks


# ── Broker client lookup ─────────────────────────────────────────────

async def _broker_client_ids(broker_id: PydanticObjectId) -> list[PydanticObjectId]:
    """Direct clients of a broker (assigned_broker_id == broker_id, trading roles)."""
    coll = User.get_motor_collection()
    cursor = coll.find(
        {
            "assigned_broker_id": broker_id,
            "role": {"$in": TRADING_ROLES},
            "status": {"$ne": UserStatus.CLOSED.value},
            "is_demo": {"$ne": True},
        },
        {"_id": 1},
    )
    return [doc["_id"] async for doc in cursor]


async def _entity_pool_ids(
    entity_id: PydanticObjectId,
    entity_role: str,
) -> list[PydanticObjectId]:
    """User IDs in an entity's pool based on role."""
    coll = User.get_motor_collection()
    query: dict[str, Any] = {
        "role": {"$in": TRADING_ROLES},
        "status": {"$ne": UserStatus.CLOSED.value},
        "is_demo": {"$ne": True},
    }
    if entity_role in (UserRole.BROKER.value, "BROKER"):
        # Whole subtree — every descendant client carries this broker's id in
        # broker_ancestry, so this captures sub-brokers' clients too (not just
        # the broker's DIRECT clients). Without it the broker's per-user list,
        # exports and sharing dropped sub-broker users (operator: "broker ko
        # uske andar ke sub-broker ke users bhi dikhe").
        query["broker_ancestry"] = entity_id
    elif entity_role in (UserRole.ADMIN.value, "ADMIN"):
        query["assigned_admin_id"] = entity_id
    elif entity_role in (UserRole.SUPER_ADMIN.value, "SUPER_ADMIN"):
        pass  # all users
    else:
        query["assigned_broker_id"] = entity_id
    cursor = coll.find(query, {"_id": 1})
    return [doc["_id"] async for doc in cursor]


# ── System-generated ADJUSTMENT filter ──────────────────────────────
#
# ADJUSTMENT is overloaded: operator-driven Add/Deduct Fund AND a
# handful of automated corrections that should NOT count as cash
# inflow or outflow on the dashboard. Each pattern below maps to a
# concrete narration the platform writes at a known site:
#
#   "Initial balance credit by …"
#       Written by user-creation flow (admin/users.py) when a new user
#       is opened with a non-zero starting balance. Onboarding entry,
#       not a real deposit.
#   "Bogus 0-price fill reversal …"
#       Written when the operator runs the fix_bogus_proceeds_credits
#       recovery script (or its inline variant) to undo a trade booked
#       at ₹0 due to a Zerodha feed flatline. The credit is a
#       correction restoring the pre-bug wallet state, not new money.
#       Operator hit this on CL62329114 — two ₹8,63,120 reversal rows
#       inflated the per-week Add Fund total to ~₹17 lakh.
#   "Reversal …" (generic)
#       Catch-all for any future correction script that prefixes its
#       narration with the word. Cheap to include; missing it would
#       just re-create the same bug for the next operator-written fix.
_SYSTEM_ADJUSTMENT_PATTERNS = (
    "initial balance credit",
    "bogus 0-price fill reversal",
    "reversal trade=",  # the bogus-fill narration always contains this fragment
)


def _is_system_adjustment(narration: str) -> bool:
    """True when an ADJUSTMENT row is a backend-written correction
    that should not be displayed as a real Add/Deduct Fund move."""
    n = (narration or "").lower()
    return any(p in n for p in _SYSTEM_ADJUSTMENT_PATTERNS)


# ── Broker totals (PNL sharing snapshot) ─────────────────────────────

async def _settlement_in_window(
    user_ids: list[PydanticObjectId],
    date_filter: dict[str, Any] | None,
) -> Decimal:
    """Settlement amount to attribute to a window.

    Without a window: returns the live `Wallet.settlement_outstanding`
    snapshot summed across the pool — same number every per-user wallet
    UI in the app shows, so lifetime reads stay consistent.

    With a window: nets SETTLEMENT_OUTSTANDING_BOOKED (accrual debits
    against the wallet when stop-out can't fully debit a loss) minus
    SETTLEMENT_OUTSTANDING_RECOVERY (auto-deducted from a later
    deposit) WalletTransaction rows whose `created_at` falls inside
    the window. The difference is the "net settlement booked in this
    period" — operator-meaningful for weekly reviews and what they
    actually want to subtract from broker P&L for a given week.
    """
    if not user_ids:
        return Decimal("0")
    if not date_filter:
        total = Decimal("0")
        wallets = await Wallet.find({"user_id": {"$in": user_ids}}).to_list()
        for w in wallets:
            total += to_decimal(w.settlement_outstanding)
        return total
    booked_q: dict[str, Any] = {
        "user_id": {"$in": user_ids},
        "transaction_type": TransactionType.SETTLEMENT_OUTSTANDING_BOOKED.value,
        "created_at": date_filter,
    }
    recovery_q: dict[str, Any] = {
        "user_id": {"$in": user_ids},
        "transaction_type": TransactionType.SETTLEMENT_OUTSTANDING_RECOVERY.value,
        "created_at": date_filter,
    }
    booked = Decimal("0")
    for t in await WalletTransaction.find(booked_q).to_list():
        booked += abs(to_decimal(t.amount))
    recovered = Decimal("0")
    for t in await WalletTransaction.find(recovery_q).to_list():
        recovered += abs(to_decimal(t.amount))
    return booked - recovered


async def compute_broker_totals(
    entity_id: PydanticObjectId,
    start_utc: datetime | None,
    end_utc: datetime | None,
) -> dict[str, Any]:
    """Compute NET CLIENT PNL, BKG, TOTAL, SETTLEMENT, ACTUAL PNL, SHARING.
    Despite the name this also serves the ADMIN and SUPER_ADMIN tabs —
    the per-role pool lookup is delegated to `_entity_pool_ids` so the
    All Users tab on /accounts-dashboard reports the admin's whole
    pool (not just direct-broker users, which were always returned by
    the legacy `_broker_client_ids` shortcut and read as 0 for admin
    callers)."""
    entity = await User.get(entity_id)
    if not entity:
        return _empty_broker_totals()

    user_ids = await _entity_pool_ids(entity_id, entity.role.value)
    # Operator-debug log — when "filter not working" complaints come in
    # we can grep journalctl for this line and confirm the dates that
    # actually reached the aggregation, ruling out a frontend bug.
    logger.info(
        "accounts_compute_broker_totals",
        extra={
            "entity_id": str(entity_id),
            "user_count": len(user_ids),
            "from_utc": start_utc.isoformat() if start_utc else None,
            "to_utc": end_utc.isoformat() if end_utc else None,
        },
    )

    net_client_pnl = Decimal("0")
    net_client_bkg = Decimal("0")
    total_deposits = Decimal("0")
    total_withdrawals = Decimal("0")

    # Build the date filter at function scope so the post-block
    # settlement helper can see it even when user_ids is empty (the
    # `_settlement_in_window` call no-ops in that case anyway, but
    # keeping the variable defined avoids a NameError).
    date_filter: dict[str, Any] | None = None
    if start_utc or end_utc:
        date_filter = {}
        if start_utc:
            date_filter["$gte"] = start_utc
        if end_utc:
            date_filter["$lte"] = end_utc

    if user_ids:
        fallback_usd_inr = to_decimal(market_data_service.get_usd_inr_rate())

        # Realized PNL — ALL positions (open with partial close PNL + closed)
        if date_filter:
            pos_query: dict[str, Any] = {
                "user_id": {"$in": user_ids},
                "status": PositionStatus.CLOSED.value,
                "closed_at": date_filter,
            }
            positions = await Position.find(pos_query).to_list()
        else:
            positions = await Position.find({"user_id": {"$in": user_ids}}).to_list()

        for p in positions:
            net_client_pnl += _realised_inr(p, fallback_usd_inr)

        # Brokerage from trades (wallet.total_brokerage is often 0 because
        # brokerage is tracked per-trade, not as a separate wallet txn)
        # Exclude trades an admin REOPEN/DELETE undid — their P&L is already
        # netted (positions deleted / reversed) and their brokerage must drop
        # out too, otherwise a deleted position keeps inflating Net Client BKG.
        bkg_trade_q: dict[str, Any] = {
            "user_id": {"$in": user_ids},
            "superseded_by_reopen": {"$ne": True},
        }
        if date_filter:
            bkg_trade_q["executed_at"] = date_filter
        bkg_trades = await Trade.find(bkg_trade_q).to_list()
        for t in bkg_trades:
            net_client_bkg += abs(to_decimal(t.brokerage))

        # Deposits in window =
        #   • user-initiated DEPOSIT requests, AND
        #   • admin manual "Add Fund" (ADJUSTMENT, amount > 0).
        # The ONLY ADJUSTMENT+ we exclude is the automatic initial_balance
        # credit written on user creation (narration "Initial balance credit
        # by …") — that's an onboarding entry, not a real deposit, and folding
        # it in once ballooned the tile to ~23 lakh for a week of signups.
        # Withdrawals mirror this with ADJUSTMENT negatives ("Deduct Fund").
        dep_query: dict[str, Any] = {
            "user_id": {"$in": user_ids},
            "transaction_type": {"$in": [TransactionType.DEPOSIT.value, TransactionType.ADJUSTMENT.value]},
        }
        if date_filter:
            dep_query["created_at"] = date_filter
        dep_txns = await WalletTransaction.find(dep_query).to_list()
        for t in dep_txns:
            amt = to_decimal(t.amount)
            if t.transaction_type == TransactionType.ADJUSTMENT:
                if amt <= 0:
                    continue  # admin Deduct Fund → counted under withdrawals
                if _is_system_adjustment(t.narration or ""):
                    # Onboarding initial balance, bogus-fill reversal, or
                    # any other auto-generated correction — NOT a real
                    # admin Add Fund. See _is_system_adjustment for the
                    # full pattern list and why it matters.
                    continue
            total_deposits += abs(amt)

        # Withdrawals in window =
        #   • user-initiated WITHDRAWAL requests, AND
        #   • admin manual "Deduct Fund" (ADJUSTMENT, amount < 0).
        wd_query: dict[str, Any] = {
            "user_id": {"$in": user_ids},
            "transaction_type": {"$in": [TransactionType.WITHDRAWAL.value, TransactionType.ADJUSTMENT.value]},
        }
        if date_filter:
            wd_query["created_at"] = date_filter
        wd_txns = await WalletTransaction.find(wd_query).to_list()
        for t in wd_txns:
            amt = to_decimal(t.amount)
            if t.transaction_type == TransactionType.ADJUSTMENT:
                if amt >= 0:
                    continue  # admin Add Fund → counted under deposits
                if _is_system_adjustment(t.narration or ""):
                    # Negative system corrections (e.g. reversal debits)
                    # are NOT real Deduct Fund actions.
                    continue
            total_withdrawals += abs(amt)

    # Settlement — windowed (BOOKED − RECOVERY WalletTransactions in
    # the selected period) when a date filter is set, full lifetime
    # wallet snapshot otherwise. Operator wants every tile on the
    # dashboard, including Settlement, to track the date picker so the
    # week-over-week comparison is honest. Lifetime fallback (no date
    # filter) keeps the wallet snapshot so the column behaves like
    # /users does when no period is chosen.
    if date_filter:
        settlement = await _settlement_in_window(user_ids, date_filter)
    else:
        settlement = Decimal("0")
        if user_ids:
            wallets = await Wallet.find({"user_id": {"$in": user_ids}}).to_list()
            for w in wallets:
                settlement += to_decimal(w.settlement_outstanding)

    # Broker view
    broker_view_pnl = -net_client_pnl
    total_of_both = broker_view_pnl + net_client_bkg
    # Actual P&L = broker take MINUS settlement. Operator-requested
    # change: previously settlement was informational only, but admins
    # were doing the subtraction by hand in every review session — bake
    # it into the headline number so the card reflects the real broker
    # economics for the window. total_of_both is still emitted so the
    # composition tile shows the pre-settlement figure for reference.
    actual_pnl = total_of_both - settlement

    # ── Broker / sub-broker sharing ──────────────────────────────────
    # Default (no formal agreement): the broker's own create/edit-form %s,
    # shown as TWO independent figures —
    #   Sharing PnL = (Total of Both − Settlement) × PnL %   (headline take)
    #   Sharing BKG = client brokerage × Brokerage %
    # (Brokerage also sits inside actual_pnl; these are display tiles, not a
    # single summed payout.) A formal ACTIVE/PAUSED PnlSharingAgreement, when
    # present, OVERRIDES this and keeps its own legacy behaviour (settlement
    # excluded) so configured agreements don't change.
    agreement = await PnlSharingAgreement.find_one({
        "broker_id": entity_id,
        "status": {"$in": [AgreementStatus.ACTIVE.value, AgreementStatus.PAUSED.value]},
    })
    agreement_type: str | None = None
    agreement_pct = to_decimal(agreement.share_pct) if agreement else Decimal("0")
    if agreement:
        agreement_type = (
            agreement.agreement_type.value
            if hasattr(agreement.agreement_type, "value")
            else str(agreement.agreement_type)
        )

    # PnL %: an ACTIVE/PAUSED agreement stays authoritative when present —
    # so agreement-based setups behave EXACTLY as before (no regression).
    # Otherwise the broker's own create-form PnL %. Brokerage %: the broker's
    # dedicated field when set, else falls back to the PnL % (pre-split
    # parity — a broker with no separate brokerage % shares it at the PnL %).
    if agreement:
        pnl_pct = agreement_pct
    else:
        pnl_pct = (
            to_decimal(entity.broker_pnl_share_pct)
            if getattr(entity, "broker_pnl_share_pct", None) is not None
            else Decimal("0")
        )
    bkg_pct = (
        to_decimal(entity.broker_brokerage_share_pct)
        if getattr(entity, "broker_brokerage_share_pct", None) is not None
        else pnl_pct
    )
    share_pct = pnl_pct
    if agreement:
        # A formal PnlSharingAgreement stays AUTHORITATIVE and unchanged — no
        # regression for setups configured on the P&L-Sharing page: P&L and
        # brokerage share on their own bases, settlement excluded.
        sharing_pnl = quantize_money(broker_view_pnl * (pnl_pct / Decimal("100")))
        sharing_bkg = quantize_money(net_client_bkg * (bkg_pct / Decimal("100")))
        if agreement.agreement_type == AgreementType.BROKERAGE_ONLY:
            sharing_pnl = Decimal("0")
    else:
        # No agreement → BOTH figures shown, exactly as the operator wants:
        #   • Sharing PnL = (Total of Both − Settlement) × PnL %   (the
        #     headline broker take on the actual P&L)
        #   • Sharing BKG = client brokerage × Brokerage %
        # These are two independent DISPLAY figures (brokerage also sits
        # inside actual_pnl) — not meant to be summed into one payout.
        sharing_pnl = quantize_money(actual_pnl * (pnl_pct / Decimal("100")))
        sharing_bkg = quantize_money(net_client_bkg * (bkg_pct / Decimal("100")))

    return {
        "net_client_pnl": str(quantize_money(net_client_pnl)),
        "net_client_bkg": str(quantize_money(net_client_bkg)),
        # Broker-view PNL = -client PNL. Emit it separately so the hero
        # card's math line in the UI can use the broker view (which is
        # what `actual_pnl` is actually computed from). Without this the
        # operator sees a math line that doesn't add up: the displayed
        # net_client_pnl is negative but the backend silently flips it.
        "broker_view_pnl": str(quantize_money(broker_view_pnl)),
        "total_of_both": str(quantize_money(total_of_both)),
        "settlement": str(quantize_money(settlement)),
        "actual_pnl": str(quantize_money(actual_pnl)),
        "sharing_pnl": str(quantize_money(sharing_pnl)),
        "sharing_bkg": str(quantize_money(sharing_bkg)),
        "total_deposits": str(quantize_money(total_deposits)),
        "total_withdrawals": str(quantize_money(total_withdrawals)),
        "share_pct": str(share_pct),
        "brokerage_share_pct": str(bkg_pct),
        "agreement_type": agreement_type,
        "client_count": len(user_ids),
    }


def _empty_broker_totals() -> dict[str, Any]:
    z = "0.00"
    return {
        "net_client_pnl": z, "net_client_bkg": z, "broker_view_pnl": z, "total_of_both": z,
        "settlement": z, "actual_pnl": z, "sharing_pnl": z, "sharing_bkg": z,
        "total_deposits": z, "total_withdrawals": z,
        "share_pct": "0", "brokerage_share_pct": "0", "agreement_type": None, "client_count": 0,
    }


# ── Per-user PNL breakdown within an entity ──────────────────────────

async def get_entity_users(
    entity_id: PydanticObjectId,
    entity_role: str,
    start_utc: datetime | None,
    end_utc: datetime | None,
    page: int = 1,
    page_size: int = 15,
    search: str | None = None,
) -> dict[str, Any]:
    """Paginated per-user PNL breakdown for an entity's client pool.

    Operator request: only ACTIVE users surface here.  Previously this
    filter was `status != CLOSED` so PENDING (never logged in) and
    BLOCKED rows polluted the table with empty 0.00 lines.  The
    aggregate pool totals card and broker tabs still walk the full
    `assigned_*` graph — this filter is local to the per-user
    pagination only.
    """
    user_query: dict[str, Any] = {
        "role": {"$in": TRADING_ROLES},
        "status": UserStatus.ACTIVE.value,
        "is_demo": {"$ne": True},
    }
    if entity_role in (UserRole.BROKER.value, "BROKER"):
        # Whole subtree — sub-brokers' clients too (broker_ancestry holds
        # this broker's id for every descendant), not just direct clients.
        user_query["broker_ancestry"] = entity_id
    elif entity_role in (UserRole.ADMIN.value, "ADMIN"):
        # Include direct users + users under admin's brokers
        broker_ids = [
            b["_id"] async for b in User.get_motor_collection().find(
                {"assigned_admin_id": entity_id, "role": UserRole.BROKER.value},
                {"_id": 1},
            )
        ]
        scope_filter = [{"assigned_admin_id": entity_id}]
        if broker_ids:
            scope_filter.append({"assigned_broker_id": {"$in": broker_ids}})
        user_query["$or"] = scope_filter

    if search and search.strip():
        import re
        escaped = re.escape(search.strip())
        search_filter = [
            {"user_code": {"$regex": escaped, "$options": "i"}},
            {"full_name": {"$regex": escaped, "$options": "i"}},
        ]
        if "$or" in user_query:
            user_query = {"$and": [user_query, {"$or": search_filter}]}
        else:
            user_query["$or"] = search_filter

    # Scan ALL users matching the scope (no DB pagination) so the
    # post-filter "drop all-zero rows" step can yield a true paginated
    # result. Earlier this used skip/limit at the DB level and then
    # filtered — so a page that fetched 10 users from Mongo would
    # display only the 3 with non-zero activity, and pagination felt
    # broken to the operator. For an admin pool (≤ a few hundred users)
    # scanning everything is fine; per-user cost is dominated by the
    # subsequent positions/trades aggregation anyway.
    users = await User.find(user_query).to_list()
    if not users:
        return {"items": [], "meta": {"page": page, "page_size": page_size, "total": 0, "total_pages": 1}}

    # Owner label per row — which broker / sub-broker / admin the client sits
    # under. Batch-resolved so the loop stays O(1) per user.
    _owner_ids = list(
        {u.assigned_broker_id for u in users if u.assigned_broker_id}
        | {u.assigned_admin_id for u in users if u.assigned_admin_id}
    )
    _owner_docs: dict[str, Any] = {}
    if _owner_ids:
        for o in await User.find({"_id": {"$in": _owner_ids}}).to_list():
            _owner_docs[str(o.id)] = o

    def _owner_of(u: Any) -> tuple[str, str]:
        if u.assigned_broker_id and str(u.assigned_broker_id) in _owner_docs:
            b = _owner_docs[str(u.assigned_broker_id)]
            kind = "Sub-broker" if (b.broker_ancestry or []) else "Broker"
            return kind, (b.full_name or b.user_code or "")
        if u.assigned_admin_id and str(u.assigned_admin_id) in _owner_docs:
            a = _owner_docs[str(u.assigned_admin_id)]
            return "Admin", (a.full_name or a.user_code or "")
        return "Direct", ""

    date_filter: dict[str, Any] = {}
    if start_utc:
        date_filter["$gte"] = start_utc
    if end_utc:
        date_filter["$lte"] = end_utc
    has_date = bool(date_filter)

    fallback_usd_inr = to_decimal(market_data_service.get_usd_inr_rate())
    user_ids = [u.id for u in users]

    # ── 5 batch $in queries instead of 417×4 per-user queries ───────
    # Previous approach (asyncio.gather or sequential loop) fired one
    # Position + Trade + Wallet + Settlement query PER USER — 1,668
    # queries for 417 users, spiking MongoDB CPU to 18-20%.  Now we
    # fetch everything in 5 bulk queries and aggregate in Python.
    if has_date:
        all_positions = await Position.find({
            "user_id": {"$in": user_ids},
            "status": PositionStatus.CLOSED.value,
            "closed_at": date_filter,
        }).to_list()
    else:
        all_positions = await Position.find({"user_id": {"$in": user_ids}}).to_list()

    # Skip reopen/delete-superseded trades — their brokerage / turnover must
    # not show in the per-user breakdown once the position is undone.
    trade_q: dict[str, Any] = {
        "user_id": {"$in": user_ids},
        "superseded_by_reopen": {"$ne": True},
    }
    if has_date:
        trade_q["executed_at"] = date_filter
    all_trades = await Trade.find(trade_q).to_list()

    all_wallets = await Wallet.find({"user_id": {"$in": user_ids}}).to_list()

    # Settlement: batch fetch booked/recovery transactions for all users
    if has_date:
        booked_txns = await WalletTransaction.find({
            "user_id": {"$in": user_ids},
            "transaction_type": TransactionType.SETTLEMENT_OUTSTANDING_BOOKED.value,
            "created_at": date_filter,
        }).to_list()
        recovery_txns = await WalletTransaction.find({
            "user_id": {"$in": user_ids},
            "transaction_type": TransactionType.SETTLEMENT_OUTSTANDING_RECOVERY.value,
            "created_at": date_filter,
        }).to_list()
    else:
        booked_txns = []
        recovery_txns = []

    # Build per-user lookup dicts from the batch results
    pos_by_user: dict = defaultdict(list)
    for p in all_positions:
        pos_by_user[p.user_id].append(p)

    trades_by_user: dict = defaultdict(list)
    for t in all_trades:
        trades_by_user[t.user_id].append(t)

    wallet_by_user: dict = {w.user_id: w for w in all_wallets}

    booked_by_user: dict = defaultdict(Decimal)
    for t in booked_txns:
        booked_by_user[t.user_id] += abs(to_decimal(t.amount))

    recovery_by_user: dict = defaultdict(Decimal)
    for t in recovery_txns:
        recovery_by_user[t.user_id] += abs(to_decimal(t.amount))

    # ── Per-user aggregation in Python (no more DB calls) ────────────
    items = []
    for u in users:
        uid = u.id
        net_pnl = sum(
            (_realised_inr(p, fallback_usd_inr) for p in pos_by_user[uid]),
            Decimal("0"),
        )
        net_bkg = sum(
            (abs(to_decimal(t.brokerage)) for t in trades_by_user[uid]),
            Decimal("0"),
        )
        total_pnl = net_pnl - net_bkg

        wallet = wallet_by_user.get(uid)
        available_bal = to_decimal(wallet.available_balance) if wallet else Decimal("0")
        if has_date:
            settlement_val = booked_by_user[uid] - recovery_by_user[uid]
        else:
            settlement_val = to_decimal(wallet.settlement_outstanding) if wallet else Decimal("0")

        if total_pnl == 0 and net_pnl == 0 and net_bkg == 0 and settlement_val == 0:
            continue

        pnl_minus_settlement = (
            total_pnl + settlement_val if total_pnl < 0 else total_pnl - settlement_val
        )
        _ok, _on = _owner_of(u)
        items.append({
            "user_id": str(u.id),
            "user_code": u.user_code or "",
            "username": u.full_name or u.user_code or "",
            "owner_kind": _ok,
            "owner_name": _on,
            "net_pnl": str(quantize_money(net_pnl)),
            "net_bkg": str(quantize_money(net_bkg)),
            "total_pnl": str(quantize_money(total_pnl)),
            "settlement": str(quantize_money(settlement_val)),
            "available_balance": str(quantize_money(available_bal)),
            "pnl_minus_settlement": str(quantize_money(pnl_minus_settlement)),
        })

    # Sort by total_pnl descending (highest first).
    items.sort(key=lambda x: float(x["total_pnl"]), reverse=True)

    # In-memory pagination on the FILTERED list so every page actually
    # contains `page_size` rows (apart from the last one). `total` now
    # reflects the count after the zero-activity filter, so the
    # frontend pagination strip matches what the eye sees.
    total_filtered = len(items)
    total_pages = max(1, (total_filtered + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    page_items = items[start:start + page_size]

    return {
        "items": page_items,
        "meta": {
            "page": page,
            "page_size": page_size,
            "total": total_filtered,
            "total_pages": total_pages,
        },
    }


async def get_all_entity_users(
    entity_id: PydanticObjectId,
    entity_role: str,
    start_utc: datetime | None,
    end_utc: datetime | None,
) -> list[dict[str, Any]]:
    """All users (no pagination) for export."""
    result = await get_entity_users(
        entity_id, entity_role, start_utc, end_utc,
        page=1, page_size=10000,
    )
    return result["items"]


# ── Excel renderers ──────────────────────────────────────────────────

_BOLD = Font(bold=True)
_GREY_FILL = PatternFill("solid", fgColor="EEEEEE")
_RIGHT = Alignment(horizontal="right")
_NUM_FMT = '#,##0.00'


def render_entity_users_excel(
    entity_name: str,
    users_data: list[dict[str, Any]],
    period_label: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PNL Report"

    ws.append([f"PNL Report — {entity_name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {period_label}"])
    ws.append([])

    headers = ["User ID", "Username", "Total PNL", "Net PNL", "Net BKG", "Settlement", "PNL - Settlement"]
    ws.append(headers)
    row_num = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = _BOLD
        cell.fill = _GREY_FILL

    for u in users_data:
        ws.append([
            u["user_code"],
            u["username"],
            float(u["total_pnl"]),
            float(u["net_pnl"]),
            float(u["net_bkg"]),
            float(u["settlement"]),
            float(u["pnl_minus_settlement"]),
        ])

    for col_idx in range(3, 8):
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = _NUM_FMT
                cell.alignment = _RIGHT

    widths = [12, 20, 15, 15, 15, 15, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data


def render_broker_totals_excel(
    totals: dict[str, Any],
    entity_name: str,
    period_label: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Broker Summary"

    ws.append([f"Broker Summary — {entity_name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {period_label}"])
    ws.append([])

    rows = [
        ("NET CLIENT PNL", totals["net_client_pnl"]),
        ("NET CLIENT BKG", totals["net_client_bkg"]),
        ("TOTAL OF BOTH", totals["total_of_both"]),
        ("SETTLEMENT", totals["settlement"]),
        ("ACTUAL PNL", totals["actual_pnl"]),
        ("SHARING PNL", totals["sharing_pnl"]),
        ("SHARING BKG", totals["sharing_bkg"]),
        ("TOTAL DEPOSITS", totals["total_deposits"]),
        ("TOTAL WITHDRAWALS", totals["total_withdrawals"]),
    ]
    for label, val in rows:
        ws.append([label, float(val)])

    for row_idx in range(4, 4 + len(rows)):
        ws.cell(row=row_idx, column=1).font = _BOLD
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.number_format = _NUM_FMT
        cell_b.alignment = _RIGHT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data


def render_single_user_excel(user_data: dict[str, Any], period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "User PNL"

    ws.append([f"PNL Report — {user_data['user_code']} ({user_data['username']})"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {period_label}"])
    ws.append([])

    rows = [
        ("Net PNL", user_data["net_pnl"]),
        ("Net BKG", user_data["net_bkg"]),
        ("Total PNL", user_data["total_pnl"]),
        ("Settlement", user_data["settlement"]),
        ("PNL - Settlement", user_data["pnl_minus_settlement"]),
    ]
    for label, val in rows:
        ws.append([label, float(val)])

    for row_idx in range(4, 4 + len(rows)):
        ws.cell(row=row_idx, column=1).font = _BOLD
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.number_format = _NUM_FMT
        cell_b.alignment = _RIGHT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data


# ── PDF renderers ────────────────────────────────────────────────────

def _money_fmt(val: str) -> str:
    try:
        f = float(val)
        sign = "+" if f > 0 else ""
        return f"{sign}{f:,.2f}"
    except (ValueError, TypeError):
        return val


def render_entity_users_pdf(
    entity_name: str,
    users_data: list[dict[str, Any]],
    period_label: str,
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"PNL Report — {entity_name}",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=14, spaceAfter=6)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements: list = []
    elements.append(Paragraph(f"PNL Report — {entity_name}", h1))
    elements.append(Paragraph(f"Period: {period_label}", meta_style))
    elements.append(Spacer(1, 8))

    headers = ["User ID", "Username", "Total PNL", "Net PNL", "Net BKG", "Settlement", "PNL-Sett."]
    data = [headers]
    for u in users_data:
        data.append([
            u["user_code"], u["username"],
            _money_fmt(u["total_pnl"]), _money_fmt(u["net_pnl"]),
            _money_fmt(u["net_bkg"]), _money_fmt(u["settlement"]),
            _money_fmt(u["pnl_minus_settlement"]),
        ])

    col_widths = [50, 70, 55, 55, 50, 50, 55]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tbl)

    doc.build(elements)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


def render_broker_totals_pdf(
    totals: dict[str, Any],
    entity_name: str,
    period_label: str,
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=f"Broker Summary — {entity_name}",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, spaceAfter=8)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements: list = []
    elements.append(Paragraph(f"Broker Summary — {entity_name}", h1))
    elements.append(Paragraph(f"Period: {period_label}", meta_style))
    elements.append(Spacer(1, 10))

    rows = [
        ["NET CLIENT PNL", _money_fmt(totals["net_client_pnl"])],
        ["NET CLIENT BKG", _money_fmt(totals["net_client_bkg"])],
        ["TOTAL OF BOTH", _money_fmt(totals["total_of_both"])],
        ["− SETTLEMENT", _money_fmt(totals["settlement"])],
        ["= ACTUAL PNL", _money_fmt(totals["actual_pnl"])],
        ["SHARING PNL", _money_fmt(totals["sharing_pnl"])],
        ["SHARING BKG", _money_fmt(totals["sharing_bkg"])],
        ["TOTAL DEPOSITS", _money_fmt(totals["total_deposits"])],
        ["TOTAL WITHDRAWALS", _money_fmt(totals["total_withdrawals"])],
    ]
    tbl = Table(rows, colWidths=[70 * mm, 60 * mm])
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#e8f5e9")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)

    doc.build(elements)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes
