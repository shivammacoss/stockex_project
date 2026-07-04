"""Renders a P&L Sharing report as an .xlsx workbook (openpyxl)."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.pnl_sharing import ReportResponse


def render_report_excel(report: ReportResponse) -> bytes:
    """Build a multi-sheet xlsx (Summary + Rows). Returns raw bytes."""
    wb = Workbook()

    # ── Summary sheet ───────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    a = report.agreement
    s = report.summary

    bold = Font(bold=True)
    grey_fill = PatternFill("solid", fgColor="EEEEEE")
    right = Alignment(horizontal="right")

    ws.append(["P&L Sharing Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Admin", a.admin_user_code or a.admin_id])
    ws.append(["Broker", a.broker_user_code or a.broker_id])
    ws.append(["Share %", a.share_pct])
    ws.append(["Mode", a.settlement_mode])
    ws.append(["Cadence", a.settlement_cadence or "-"])
    ws.append(["Status", a.status])
    ws.append([])
    ws.append(["Total Sharing PNL", s.total_sharing_pnl_inr])
    ws.append(["Total Sharing BKG", s.total_sharing_bkg_inr])
    ws.append(["Periods Settled", s.periods_settled])
    ws.append(["Periods Pending", s.periods_pending])
    ws.append(["Periods Failed", s.periods_failed])
    ws.append(["Periods Unsettled", s.periods_unsettled])
    for row_idx in range(3, 16):
        ws[f"A{row_idx}"].font = bold
        ws[f"B{row_idx}"].alignment = right
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    # ── Rows sheet ──────────────────────────────────────────
    rows_ws = wb.create_sheet("Rows")
    headers = [
        "Period start",
        "Period end",
        "Net Client PNL",
        "Net Client BKG",
        "Total of both",
        "Actual PNL",
        "Sharing PNL",
        "Sharing BKG",
        "Status",
    ]
    rows_ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = rows_ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = grey_fill
    for r in report.rows:
        rows_ws.append(
            [
                r.period_start.strftime("%Y-%m-%d %H:%M"),
                r.period_end.strftime("%Y-%m-%d %H:%M"),
                r.net_client_pnl_inr,
                r.net_client_bkg_inr,
                r.total_of_both_inr,
                r.actual_pnl_inr,
                r.sharing_pnl_inr,
                r.sharing_bkg_inr,
                r.settlement_status,
            ]
        )
    for col_idx in range(1, len(headers) + 1):
        rows_ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    buf.close()
    return excel_bytes
